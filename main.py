from encodings import utf_8
from selenium import webdriver
from selenium.common import NoSuchElementException, ElementClickInterceptedException, ElementNotInteractableException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import pandas as pd
from openpyxl import load_workbook
import time

encoding: utf_8
# filename = 'Naturgy_datos.xlsx'
# filename = 'SEBAS.xlsx'
filename = 'Data_Prueba.xlsx'
filesheet = pd.read_excel(filename, usecols='E:H')

workbook = load_workbook(filename)
worksheet = workbook.active

cp = list(filesheet.loc[:, 'CP'])
dir = list(filesheet.loc[:, 'DIRECCION'])
loc = list(filesheet.loc[:, 'LOCALIDAD'])
cp_valor_final = []

for i in range(len(cp)):
    if not pd.isna(cp[i]):
        cp_valor_final.append(str(int(cp[i])).zfill(5))

valor_cup = ""
valor_potenciaP1 = ""
cups = []
potenciasP1 = []


class webScrapping():
    def recolectar_cups_potenciaP1(self):
        for i in range(len(cp)):
            try:
                # DAR CLICK EN DIRECCION
                driver.find_element(By.XPATH, "//a[normalize-space()='introducir tu dirección']").click()
                time.sleep(5)

            except ElementClickInterceptedException:
                driver.refresh()
                continue
            # Poner la direccion
            try:
                direccion = driver.find_element(By.XPATH,
                                                "//section[@role='main']//div//div//div//div//div//div//div//div//div//div//div//div//div//div//div//input[@role='combobox']")
                direccion.send_keys(dir[i])
                time.sleep(10)
                direccion.send_keys(Keys.DOWN)
                direccion.send_keys(Keys.ENTER)
            except NoSuchElementException:
                driver.refresh()
                continue
            try:
                time.sleep(5)
                piso = driver.find_element(By.XPATH,
                                           "//body/div/section[@role='main']/div/div/div/div/div/div/div/div/div/div/div/div/div/div/input[1]")
                piso.click()
                time.sleep(5)
                piso.send_keys(Keys.DOWN)
                piso.send_keys(Keys.ENTER)
                time.sleep(10)
            except ElementNotInteractableException:
                driver.refresh()
                continue
            try:
                driver.find_element(By.XPATH,"//body[1]/div[1]/section[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[4]/div[1]/button[1]").click()
            except NoSuchElementException:
                driver.refresh()
                continue
            time.sleep(10)

            try:
                if not pd.isna(loc[i]):
                    # Elegir la localidad
                    localidad = driver.find_element(By.CSS_SELECTOR, "#city")
                    selector_localidad = Select(localidad)
                    selector_localidad.select_by_visible_text(loc[i])
                else:
                    continue
            except NoSuchElementException:
                driver.refresh()
                continue
            try:
                # Poner el valor de la calle
                calle = driver.find_element(By.CSS_SELECTOR, "#street")
                calle.send_keys(dir[i])
                time.sleep(15)
                calle.send_keys(Keys.DOWN)
                calle.send_keys(Keys.ENTER)
            except NoSuchElementException:
                driver.refresh()
                continue

            # Darle al boton de calcular
            driver.find_element(By.XPATH, "(//button[@type='submit'])[2]").click()
            time.sleep(30)

            try:
                # Obtiene el cup y la potenciaP1
                cup = driver.find_element(By.XPATH,
                                          "//body/div[@id='root']/section[@role='main']/div/div/div/div/div[2]/p[1]")
                potenciaP1 = driver.find_element(By.XPATH, "//p[3]")
            except NoSuchElementException:
                driver.refresh()
                continue

            valor_cup = cup.text[6:len(cup.text)]
            cups.append(valor_cup)

            valor_potenciaP1 = potenciaP1.text[24:len(potenciaP1.text)]
            potenciasP1.append(valor_potenciaP1)

            for a, value in enumerate(cups):
                worksheet.cell(row=i + 2, column=10, value=value)
            for a, value in enumerate(potenciasP1):
                worksheet.cell(row=i + 2, column=11, value=value)

            driver.refresh()


url = 'https://front-calculator.zapotek.adn.naturgy.com/'
driver = webdriver.Chrome()
driver.maximize_window()
driver.get(url)
trabajo = webScrapping()
trabajo.recolectar_cups_potenciaP1()

workbook.save(filename)
